from schemas import AdminInterface
import openpyxl

# Глобальные переменные
excel_file = 'admin/Menu.xlsx'

# Определение экземпляров классов
update = AdminInterface()

book = ''
data = ''


class Engine:

    def start(self):
        global book
        global data
        book = openpyxl.open(excel_file)
        data = book.active

        for row in data.iter_rows(min_col=1, max_col=3):
            if row[0].value != None and row[1].value != None and row[2].value != None:
                update.menu_data(row[0].value, row[1].value, row[2].value)
                self.submenus_add(row[0].value)

    def submenus_add(self, id: str):
        for row in data.iter_rows(min_col=1, max_col=4):
            if row[0].value == id:

                for cell in data.iter_rows(min_row=row[0].row):
                    if int(cell[0].value or 0) <= id:
                        if cell[1].value != None and cell[2].value != None and cell[3].value != None:
                            update.submenu_data(
                                cell[1].value, id, cell[2].value, cell[3].value)
                            self.dish_add(id, cell[1].value)
                    else:
                        break

    def dish_add(self, menu_id: str, submenu_id: str):
        for row in data.iter_rows():
            if row[0].value == menu_id:

                for cell in data.iter_rows(min_row=row[0].row+1, min_col=2):
                    if cell[0].value == submenu_id:

                        for sub_cell in data.iter_rows(min_row=cell[0].row+1, min_col=3):
                            if sub_cell[0].value != None and sub_cell[1].value != None and sub_cell[2].value != None and sub_cell[3].value != None:
                                update.dish_data(
                                    menu_id, submenu_id, sub_cell[0].value, sub_cell[1].value, sub_cell[2].value, sub_cell[3].value)
                            else:
                                break
